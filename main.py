import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import argparse


def get_cell_name(ws: Worksheet, row: int, column: int) -> str:
    return ws.cell(row, column).coordinate

def set_cell_value(ws: Worksheet, row: int, column: int, value):
    ws.cell(row, column).value = value


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('-p', '--players', type=str, required=True, help='Players who will attend the game.')
    args = parser.parse_args()

    filename = "/mnt/c/Users/danie/Desktop/output.xlsx"
    max_rounds = 1000

    wb = openpyxl.Workbook()
    ws = wb.active

    players = str(args.players).split()

    for i, player in enumerate(players):
        col_index = i * 4 + 1
    
        ws.merge_cells(start_row=1, start_column=col_index, end_row=1, end_column=col_index+1)

        set_cell_value(ws, 1, col_index, player)
        set_cell_value(ws, 1, col_index+2, "WYNIK")
        set_cell_value(ws, 1, col_index+3, f"=SUM({get_cell_name(ws, 3, col_index)}:{get_cell_name(ws, 3, col_index+2)})")

        set_cell_value(ws, 2, col_index, "Rzut 1")
        set_cell_value(ws, 2, col_index+1, "Rzut 2")
        set_cell_value(ws, 2, col_index+2, "Rzut 3")

        ws.merge_cells(start_row=2, start_column=col_index+3, end_row=3, end_column=col_index+3)
        set_cell_value(ws, 2, col_index+3, "SUMA RUNDY")

        set_cell_value(ws, 3, col_index, f"=SUM({get_cell_name(ws, 4, col_index)}:{get_cell_name(ws, 4 + max_rounds, col_index)})")
        set_cell_value(ws, 3, col_index+1, f"=SUM({get_cell_name(ws, 4, col_index+1)}:{get_cell_name(ws, 4 + max_rounds, col_index+1)})")
        set_cell_value(ws, 3, col_index+2, f"=SUM({get_cell_name(ws, 4, col_index+2)}:{get_cell_name(ws, 4 + max_rounds, col_index+2)})")

    wb.save(filename)

if __name__ == "__main__":
    main()



